"""
Read Chamber run log R3.xlsx and look up run numbers for uniformity SP runs.

Structure of the workbook
─────────────────────────
One worksheet per chamber (V1 … V7).
Each row: Date | SO | Run Number | Customer | Recipe | Notes | ...

The run number column contains IDs such as:
  V2-1986, V2-unif, V2 Hf Unif, V7-HW, V1-SiO2 Uniformity, …

Matching strategy
─────────────────
Given an SP anchor file with (chamber, date_str, design_name):
 1. Find all run log rows for that chamber within ±1 day of date_str.
 2. Among those, rank by word-overlap between design_name and Recipe.
 3. Return the Run Number of the best-scoring match (or None if no match).

Results are cached in memory after the first load.
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Optional

import openpyxl

_DEFAULT_LOG = (
    Path(r"C:\Users\User\FiveNine Dropbox\FiveNine Optics Team Folder"
         r"\Coating Run Data\Chamber run log R3.xlsx")
)

_DATE_FMT = "%m/%d/%Y"


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _words(text: str) -> set[str]:
    """Lower-cased word tokens, stripping punctuation."""
    return set(re.findall(r"[a-z0-9]+", text.lower()))


def _overlap(a: str, b: str) -> int:
    """Number of shared word tokens between two strings."""
    return len(_words(a) & _words(b))


def _parse_date(val) -> Optional[datetime]:
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                pass
    return None


# ---------------------------------------------------------------------------
# RunLogReader
# ---------------------------------------------------------------------------

class RunLogReader:
    """Load and index a chamber run log workbook.

    Usage:
        reader = RunLogReader()                # default path from CLAUDE.md
        rn = reader.find_run_number("V2", "06/02/2026", "Hf layer 350nm")
        # → "V2 Hf Unif" or None
    """

    def __init__(self, log_path: str | Path | None = None):
        self._path  = Path(log_path) if log_path else _DEFAULT_LOG
        self._index: dict[str, list[tuple[datetime, str, str]]] = {}
        # index structure: {chamber: [(date, run_number, recipe), ...]}
        self._loaded = False

    def _ensure_loaded(self) -> None:
        if self._loaded:
            return
        if not self._path.exists():
            return
        wb = openpyxl.load_workbook(
            str(self._path), read_only=True, data_only=True
        )
        for ch in ["V1", "V2", "V3", "V4", "V5", "V6", "V7"]:
            if ch not in wb.sheetnames:
                continue
            ws = wb[ch]
            if not hasattr(ws, "iter_rows"):
                continue
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # header
                date = _parse_date(row[0])
                run_number = str(row[2]).strip() if row[2] is not None else ""
                recipe     = str(row[4]).strip() if row[4] is not None else ""
                if date and run_number:
                    rows.append((date, run_number, recipe))
            self._index[ch] = rows
        wb.close()
        self._loaded = True

    def find_run_entry(
        self,
        chamber: str,
        date_str: str,
        design_name: str,
        window_days: int = 7,
    ) -> Optional[tuple[str, str]]:
        """Return (run_number, run_log_date_str) for the best match, or None.

        Searches ±window_days around date_str (the SP measurement date) to
        account for the delay between coating and spectrophotometer measurement.
        Candidates are ranked by word-overlap between design_name and Recipe.

        Returns the run log date (coating date) not the SP measurement date,
        so the dashboard tracks when the coating was actually done.
        """
        self._ensure_loaded()
        rows = self._index.get(chamber, [])
        if not rows:
            return None

        try:
            target = datetime.strptime(date_str, _DATE_FMT)
        except (ValueError, TypeError):
            return None

        window = timedelta(days=window_days)
        candidates = [
            (run_number, recipe, date)
            for (date, run_number, recipe) in rows
            if abs(date - target) <= window
        ]
        if not candidates:
            return None

        # Rank by word overlap with the SP design name; fall back to first
        ranked = sorted(
            candidates,
            key=lambda t: _overlap(design_name, t[1]),
            reverse=True,
        )
        run_number, _, log_date = ranked[0]
        return run_number, log_date.strftime(_DATE_FMT)

    def find_run_number(
        self,
        chamber: str,
        date_str: str,
        design_name: str,
        window_days: int = 7,
    ) -> Optional[str]:
        """Convenience wrapper — returns just the run number (or None)."""
        entry = self.find_run_entry(chamber, date_str, design_name, window_days)
        return entry[0] if entry else None

    def is_available(self) -> bool:
        """True if the run log file can be read."""
        return self._path.exists()
